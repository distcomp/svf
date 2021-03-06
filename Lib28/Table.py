# -*- coding: cp1251 -*-

from   numpy import *
import   numpy as np
from   os.path  import *
import openpyxl
import sys
import COMMON as co
from   Pars   import *
##from   InData import *


def  TblFldV ( TblFld ) : return co.Task.getTbl_tbl ( TblFld )
def  Tbl     ( nTbl ) :   return co.Task.getTbl     ( nTbl   )


def joinTab(*Tabs):
        for itbl, tbl in enumerate(Tabs):
            if itbl == 0:
                ret = deepcopy(tbl)
            else:
                for fld in tbl.Flds:
                    if ret.getFieldNum(fld.name) == -1:
                        ret.Flds.append(deepcopy(fld))
            print(tbl.NoR)

        co.Task.KillTbl(ret.name)  # kill the same name
        co.Task.AddTbl(ret)  # ��� ������ �������
        co.curentTabl = ret
        return ret


def joinTabBy(By, *Tabs):
    for j, tbl in enumerate(Tabs):
        if j == 0:
            ret = deepcopy(tbl)
            retByNum = ret.getFieldNum(By)
            if retByNum == -1:
                print ( 'Err find Field by', By )
                exit (-1)
        else:
            tbj = deepcopy(tbl)
            tbjByNum = tbj.getFieldNum(By)
            if tbjByNum == -1:
                print ( 'Err find Field by', By )
                exit (-1)

            r = -1
            print ('shape', ret.Flds[retByNum].tb.shape[0])
            while ( r < ret.NoR-1 and r < tbj.NoR-1 ):
                r +=1
                print (r, ret.Flds[retByNum].tb[r] , tbj.Flds[tbjByNum].tb[r])
                if ret.Flds[retByNum].tb[r] == tbj.Flds[tbjByNum].tb[r] : continue
                ind = where (tbj.Flds[tbjByNum].tb[r:]==ret.Flds[retByNum].tb[r])
#                print (ind[0])
 #               print (len(ind[0]))
  #              print (ind[0].shape)
                if ind[0].size == 0 :
                    ret.KillRow (r)
                    r -= 1
                    continue
                else :
                    tbj.KillRow (r)
                    r -= 1
                    continue

            while ret.NoR != tbj.NoR :
                if ret.NoR > tbj.NoR : ret.KillRow (ret.NoR-1)
                else                 : tbj.KillRow (tbj.NoR-1)
            print('NoR RET TBJ', ret.NoR, tbj.NoR)
            ret = joinTab (ret, tbj)


    co.Task.KillTbl(ret.name)  # kill the same name
    co.Task.AddTbl(ret)  # ��� ������ �������
    co.curentTabl = ret
    return ret


def appendTab(*Tabs):
        for itbl, tbl in enumerate(Tabs):
            if itbl == 0:
                ret = deepcopy(tbl)
            else:
                for ifld, rfld in enumerate(ret.Flds):
                    rfld.tb = append (rfld.tb, tbl.Flds[ifld].tb)
        ret.NoR = ret.Flds[0].tb.size
#        print(ret.NoR)
        co.Task.KillTbl(ret.name)  # kill the same name
        co.Task.AddTbl(ret)  # ��� ������ �������
        co.curentTabl = ret
        return ret


def  Select ( buf ):
                buf = buf.strip()
                part = buf.split( ' from ' )
                fields = part[0]
                part = part[1].split( ' where ' )
                fromFile = addDataPath ( part[0] )
                if len (part) == 1 :                    #  no where
                    where = ''
                else :
                    where = co.Task.substitudeDef(part[1])
                print (fields, '\n', fromFile, '\n', where)
                return Table ( fromFile, fields, where )



def  arrayLatLonToAzimut(lat, lon, x, y) :
        for i in range(x.shape[0]):  x[i],y[i] = LatLonToAzimut( lat[i],lon[i] )


def  TblOperation (buf):
        if buf == '' : return
        buf = co.Task.substitudeDef(buf)
        buf = buf.replace(' ', '') 
        print  (buf)
        parts = buf.split ( '=TblLatLonToAzimut(' )             #  LatLonToAzimut
        if len(parts)==1 : parts = buf.split ( '=TblLatLonToGaussKruger(' )    # LanLonToGaussKruger
        if len(parts)==2 :
            parts[1] = parts[1].split(')')[0]
            if co.printL: print (parts)
            lefts = parts[0].split(',')
            x_tbl = co.Task.getTbl_tbl (lefts[0])
            y_tbl = co.Task.getTbl_tbl (lefts[1])
            rights = parts[1].split(',')
            lat_tbl = co.Task.getTbl_tbl (rights[0])
            lon_tbl = co.Task.getTbl_tbl (rights[1])
            if buf.find ('TblLatLonToAzimut') > 0 :
                for i in range(x_tbl.shape[0]):  x_tbl[i],y_tbl[i] = LatLonToAzimut( lat_tbl[i],lon_tbl[i] )
            else :
                for i in range(x_tbl.shape[0]):  x_tbl[i],y_tbl[i] = WGS84toGausKru( lat_tbl[i],lon_tbl[i],0 )
            return

        parts = buf.split ( '.AddField(' )             #  AddField
        if len(parts)==2 :
            parts[1] = parts[1].split(')')[0]
            co.Task.getTbl(parts[0]).AddField( parts[1] )
            print ('AddField :', parts[1], 'to', parts[0])
            return
        
        parts = buf.replace ( ' ', '').split ( '=' )
        left  = parts[0]
        right = parts[1]
        for itb, tb in enumerate ( co.Task.Tbls ) :
#            print 'tb.name', tb.name
            for ifld, fld in enumerate (tb.Flds) :
                right = right.replace ( tb.name +'.'+ fld.name,
                                        'com.Task.Tbls['+str(itb)+'].Flds['+str(ifld)+'].tb')
        print (right)

#        tb = com.Task.getTbl_tbl(left)
        tb_num, fld_num = com.Task.getTblFieldNums (left)
        com.Task.Tbls[tb_num].Flds[fld_num].tb = eval (right)
        print ('Operation', parts)
#        1/0
        return
       

class Field :
    def __init__ ( self, name, src_name ): 
        self.name     = name
        self.tb       = None
        self.src_name = src_name
        self.src_num  = NaN
        
    def Mprint ( self ) :  print ('Field :', self.name, self.src_name, self.src_num)



class Table :
    def __init__ ( self, fromFile, fields='*', where='' ): #, useNaN=False ) :
        self.name      = ''
        self.FileType  = 'tbl' 
        self.FileVer   = 0
        self.Flds      = []
        
###        self.cols    = []
###        self.nums    = []
        self.NoC     = 0
        self.NoR     = 0
        self.sR      = []
        self.useNaN  = co.useNaN

        if co.printL : print ('\nSelect', fields, '\n  from', fromFile,'\n   where', where)

        ff_nn = SplitIgnor ( fromFile, ' AS ' )     # ��� ����� � �������
        self.fromFile = ff_nn[0]
        if len (ff_nn) > 1 : self.name = ff_nn[1]
        else               : self.name = 'curentTabl'

        _fields = fields.split(',')                   # Fields
        if co.printL : print (_fields)
        for fld in _fields :
            part = SplitIgnor ( fld.strip(), ' AS ' )
            src_name = part[0]
            if len(part) == 2 : name = part[1]              # As
            else              : name = part[0]              # the same name
            self.Flds.append ( Field ( name, src_name ) )
            if src_name.upper() == 'ROWNUM' :  self.Flds[-1].src_num = -1
            if co.printL : self.Flds[-1].Mprint()

        root, ext = splitext(self.fromFile.upper())
        if   co.Task.getTblNum(self.fromFile) !=-1 :  self.Read21_TBL ( where )
        elif co.Task.getFunNum(self.fromFile) !=-1 :  self.Read28_FUN ( where )
        elif   '.XLSX' == ext :                       self.Read21_XLSX( where )
        elif   '.KML'  == ext :                       self.Read21_KML ( where )
        elif   '.ASC'  == ext :                       self.Read27_ASC ( where )
        else :                                        self.Read25_TXT ( where )

        print ('NoR =', self.NoR, '\n')
        if co.printL :
            for ifld, fld in enumerate (self.Flds) :
                print ('Field :', fld.name, fld.src_name, fld.src_num, fld.tb.min(0), fld.tb.max(0))

        self.sR = range (self.NoR)

        co.Task.KillTbl ( self.name )   # kill the same name
        co.Task.AddTbl ( self )
        co.curentTabl = self 


    def AddField( self, name ) :
        self.Flds.append ( Field ( name, '' ) )
        self.Flds[-1].tb = zeros ( (self.NoR), float64 )


    def IndexCol( self, Name ) :
        return self.getFieldNum(Name)

    def getFieldNum (self, name) :
        for ifi, ofi in enumerate (self.Flds) :
            if ofi.name == name:  return ifi
        return -1

    def getField (self, name) :
        for ofi in self.Flds:
            if ofi.name == name:   return ofi
        return None

    def getField_tb (self, name) :
        for ofi in self.Flds:
            if ofi.name == name:   return ofi.tb
        return None

    def dat (self, name) :              #  the same
        for ofi in self.Flds:
            if ofi.name == name:   return ofi.tb
        return None


    def KillField (self, name) :
                ind = self.getFieldNum (name) 
                if ind >= 0 : del  self.Flds[ind]

    def KillRow (self, num) :
        for ofi in self.Flds:
           ofi.tb = delete(ofi.tb, num)
        self.NoR -= 1

    def Operation ( self, buf ) :
        buf = co.Task.substitudeDef(buf)
        part = buf.replace ( ' ', '').split ( '=' )
        print ('Operation', part)
        rightPart = part[1]
        leftCol = part[0].split('.')[1]
###        leftInd = self.cols.index (leftCol)
        leftInd = self.getFieldNum(leftCol)
#        print 'leftPart', leftCol, leftInd
###        for ic, c in enumerate(self.cols) :
   ###         rightPart = rightPart.replace('curentTabl.'+c, 'self.tbl[:,'+str(ic)+']')
        for ic, c in enumerate(self.Flds) :
            rightPart = rightPart.replace('curentTabl.'+c.name, 'c.tb[:]')
        print (rightPart)
#        print self.tbl[:,1]
###        self.tbl[:,leftInd] = eval (rightPart)
        self.Flds[leftInd].tb[:] = eval(rightPart)
#        print self.tbl[:,1]
#        1/0

    def TblLonLatToGaussKruger ( self ) :
            print ('col 0 and 1 convert To GaussKruger')
#            print 'degree :', self.tbl[0][0], self.tbl[0][1]
            for i in range(self.NoR):
                self.Flds[0].tb[i], self.Flds[1].tb[i] = WGS84toGausKru(self.Flds[1].tb[i], self.Flds[0].tb[i], 0)
###            for i in range(self.tbl.shape[0]):
   ###             self.tbl[i][0], self.tbl[i][1] = WGS84toGausKru(self.tbl[i][1], self.tbl[i][0], 0)
#            print 'GaussKruger :', self.tbl[0][0], self.tbl[0][1]


    def Read21_XLSX ( self, where ) :
        try :
            wb = openpyxl.load_workbook(self.fromFile)
            ws = wb[wb.get_sheet_names()[0]]
        except :
            print ('No file ', self.fromFile)
            exit (-1)
        else:
            print ('max column/row', ws.max_column, ws.max_row,)
            for nameRow in range (1,1000):           # Looking for NANEs string
                if ws.cell(row=nameRow,column=1).value == None : continue
                if ws.cell(row=nameRow,column=1).value[0] == '#' : continue
                break
            print ('nameRow', nameRow)
            names = []
            for col in range(1,ws.max_column+1) :           # co - number of collomns - till first None
                name = str (ws.cell(row=nameRow,column=col).value)
                if name == '#END#' : break
                if name == None : name = ''
#                print name
                names.append(name.strip())           # remuve end  blancs
            NoC = len (names)
            if co.printL :  print ("TablesNames", names)

            print ("TablesNames", names)
            for fld in self.Flds :
                if isnan(fld.src_num) :              # not  ROWNUM
                  try :
                    fld.src_num = names.index ( fld.src_name )
                  except :    
                    print ("No Column for", fld.src_name, "*****************************")
                    exit (-1)
                fld.Mprint()    

###            for fld in self.Flds :  self.nums.append( fld.src_num )    #  Kill
###            for fld in self.Flds :  self.cols.append( fld.name )    #  Kill
                    
            self.NoC = len(self.Flds)  #(self.cols)
            maxNoR = ws.max_row  #50000
            for fld in self.Flds : fld.tb = zeros ( maxNoR, float64 ) 

            NoR = 0;
            ro = nameRow
            for t_raw in range (ws.max_row-nameRow)  :  #while (1) :
                ro += 1
#                print ro, ws.cell(row=ro,column=1).value
 #               if ws.cell(row=ro,column=1).value == None  : break
                try :
                    if str(ws.cell(row=ro,column=1).value) == '#END#' : break
                    if str(ws.cell(row=ro,column=1).value)[0] == '#'  : continue
                except ValueError:  pass
 #                   print '***********Row ommited***********  rownum=', ro, ws.cell(row=ro,column=1).value
  #                  continue
                wher = where[0:]
                OK = True
                for fld in self.Flds :
  #                  print (pceil.value, pceil.data_type )
                    if fld.src_num == -1 :
                        fld.tb[NoR] = NoR
                        continue
                    pceil = ws.cell(row=ro, column=fld.src_num + 1)
#                    print (com.TaskName[:5])
                    if com.TaskName[:5] == 'COVID' and fld.name == 'iso_code' :     ############################  COVID
 #                       print (pceil.value)
                        fld.tb[NoR] = strTOnum(str(pceil.value))
#                    elif com.TaskName[:5] == 'COVID' and fld.name == 'date' :     ############################  COVID
#                        fld.tb[NoR] = float(str(pceil.value).replace('-',''))
                    else:
                        fld.tb[NoR] = floatGradNaN(ws.cell(row=ro,column=fld.src_num+1).value)
#                        try :
 #                           fld.tb[NoR] = float ( ws.cell(row=ro,column=fld.src_num+1).value )
  #                      except :
   #                         fld.tb[NoR] = NaN
                    if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
#                    print (fld.tb[NoR])
                if not OK: continue
                if len (where) >= 3 :                       #check  where
                    wher = where.replace('ROWNUM',str(NoR))
                    for fld in self.Flds :   wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
 #                   if self.fromFile == '../Moscow.xlsx':
  #                    print ('wher:', wher, NoR, ws.max_row)
                    if eval (wher, {'strTOnum': strTOnum}) == False : continue
                NoR += 1
                if NoR >= maxNoR :                          # resize
                    maxNoR *= 2
                    print ('resize to',  maxNoR)
                    for fld in self.Flds : fld.tb = resize (fld.tb, maxNoR ) 
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR ) 
            
            self.NoR = NoR

    def setField_src_num( self, names ):        # ������������ � �� �������� !
            for fld in self.Flds:
                if fld.src_name == '*':
                    for na in names: self.Flds.append(Field(na, na))
                    self.Flds.append(Field('ROWNUM', 'ROWNUM'))
                elif fld.src_name == 'ROWNUM':   fld.src_num = -1
                else:
                    fld.src_num = names.index(fld.src_name)
                    if fld.src_num == -1:
                        print("No Column for", fld.src_name, "*****************************")
                        exit(-1)
            self.KillField('*')
            self.NoC = len(self.Flds)

    def Read28_FUN(self, where ) :
            srcFun = co.Task.getFun ( self.fromFile )
            names = []                                      # ����� ���������� � �������
            for a in srcFun.A:   names.append ( a.name )
            names.append(srcFun.V.name)

            self.setField_src_num ( names )

            for fld in self.Flds : fld.Mprint()
            srcNoR = 1
            for a in srcFun.A: srcNoR *= (a.Ub + 1)
            for fld in self.Flds : fld.tb = zeros ( srcNoR, float64 )

            NoR = 0

            if srcFun.dim == 1:
                Ax = srcFun.A[0]
#                print (Ax.name)
                for x in Ax.NodS:
                    OK = True
                    for fld in self.Flds :
                        if   fld.src_num==-1 :  fld.tb[NoR] = NoR
                        elif fld.src_num== 0 :  fld.tb[NoR] = Ax.min + Ax.step * x
                        else :
                            fld.tb[NoR] = srcFun.grdNaNreal(r)
                            if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break
                    if not OK: continue
                    if len (where) >= 3 :                       #check  where
                        wher = where.replace('ROWNUM',str(NoR))
                        for fld in self.Flds :
                            wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
                        if eval (wher,{}) == False : continue
                    NoR += 1
            elif srcFun.dim == 2:
                Ax = srcFun.A[0]
                Ay = srcFun.A[1]
                for y in Ay.NodS:
                  for x in Ax.NodS:
                    OK = True
                    for fld in self.Flds :
                        if   fld.src_num==-1 :  fld.tb[NoR] = NoR
                        elif fld.src_num== 0 :  fld.tb[NoR] = Ax.min + Ax.step * x
                        elif fld.src_num== 1 :  fld.tb[NoR] = Ay.min + Ay.step * y
                        else :
                            fld.tb[NoR] = srcFun.grdNaNreal(x,y)
                            if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break
                    if not OK: continue
                    if len (where) >= 3 :                       #check  where
                        wher = where.replace('ROWNUM',str(NoR))
                        for fld in self.Flds :
                            wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
                        if eval (wher,{}) == False : continue
                    NoR += 1
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR )
            self.NoR = NoR


    def Read21_TBL(self, where ) :    #  tbl -> fld.tb
            source_tb = co.Task.getTbl ( self.fromFile )
            srcFlds = source_tb.Flds
            for fld in self.Flds :
                if fld.src_name == '*' :
                    for s_fld in srcFlds : self.Flds.append ( Field ( s_fld.name, s_fld.name ) )
                elif fld.src_name == 'ROWNUM' : fld.src_num = -1
                else :    
                    fld.src_num = source_tb.getFieldNum (fld.src_name)
                    if fld.src_num == -1 :
                        print ("No Column for", fld.src_name, "*****************************")
                        exit (-1)    
            self.KillField ('*')
            for fld in self.Flds : fld.Mprint()

            self.NoC = len(self.Flds)
            for fld in self.Flds : fld.tb = zeros ( source_tb.NoR, float64 ) 

            NoR = 0
            for r in range (source_tb.NoR) :
                wher = where[0:]
                OK = True
                for c, fld in enumerate(self.Flds) :
                    if fld.src_num==-1 :  fld.tb[NoR] = NoR
                    else:                 fld.tb[NoR] = source_tb.Flds[fld.src_num].tb[r]
                    if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
                if not OK: continue
                if len (where) >= 3 :                       #check  where
                    wher = where.replace('ROWNUM',str(NoR))
                    for c , fld in enumerate(self.Flds) :
                        wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
                    if eval (wher,{'strTOnum': strTOnum}) == False : continue
                NoR += 1
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR )
            self.NoR = NoR




    def Read21_KML(self, where ) :   
        srcFlds = []                                            # ������� - ��������� �������
        srcFlds.append ( Field ( 'X', 'X' ) );  srcFlds[0].tb = [] 
        srcFlds.append ( Field ( 'Y', 'Y' ) );  srcFlds[1].tb = []

        with open( self.fromFile, "r") as fi:
            for line in fi :
                if line.find('<coordinates>') != -1 : break
            for line in fi :
                if line.find('</coordinates>') != -1 : break
                points = line.split()
                print (points)
                for p in points :
                    xyz = p.split(',')
                    srcFlds[0].tb.append ( float(xyz[0]) )       
                    srcFlds[1].tb.append ( float(xyz[1]) )       
                    print (xyz)
            srcNoR = len(srcFlds[0].tb)
#            source_tb = co.Task.getTbl ( self.fromFile )
            for fld in self.Flds :
                if fld.src_name == '*' :
                    for s_fld in srcFlds : self.Flds.append ( Field ( s_fld.name, s_fld.name ) )
                elif fld.src_name == 'ROWNUM' : fld.src_num = -1
                else :    
#                    fld.src_num = source_tb.getFieldNum (fld.src_name)
                    for is_fld, s_fld in enumerate(srcFlds) :
                         if s_fld.name == fld.src_name :
                             fld.src_num = is_fld
                             break
                    if fld.src_num == -1 :
                        print ("No Column for", fld.src_name, "*****************************")
                        exit (-1)    
            self.KillField ('*')
            for fld in self.Flds : fld.Mprint()

                        # To Kill
###            for fld in self.Flds :
   ###             self.cols.append (fld.name)
 ###               self.nums.append (fld.src_num)
      ###      print '***', self.cols ###, self.nums
                
            self.NoC = len(self.Flds)
            for fld in self.Flds : fld.tb = zeros ( srcNoR, float64 ) 

            NoR = 0;
            for r in range (srcNoR) :
                wher = where[0:]
                OK = True
                for c, fld in enumerate(self.Flds) :
                    if fld.src_num==-1 :  fld.tb[NoR] = NoR
                    else:                 fld.tb[NoR] = srcFlds[fld.src_num].tb[r]
                    if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
                if not OK: continue
                if len (where) >= 3 :                       #check  where
                    wher = where.replace('ROWNUM',str(NoR))
                    for c , fld in enumerate(self.Flds) :
                        wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
                    if eval (wher,{}) == False : continue    
                NoR += 1
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR )
            self.NoR = NoR


    def Read25_TXT( self, where ) :      #    tbl -> fld.tb
        if co.printL: print ('Read25_TXT')
        with open( self.fromFile, "r") as fi:
            line1 = fi.readline()
            names_ver_type = line1.split('#SvFver_')      # Version & Type of File
            if len (names_ver_type) ==2 :
                parts = names_ver_type[1].split('_')
                self.FileVer  = int(parts[0])
                self.FileType = parts[1].split()[0]       # ������ ����. ������
            if co.printL : print ("FileVerType", self.FileVer, self.FileType)
            
            names = names_ver_type[0].split()
            if self.FileVer==0 and isfloat(names[0]) :       # if ��� ����� � �� ��� - ������� ��� ���� ��������
                names = ['Col'+str(c) for c in range(len(names)) ]
                fi.seek(0)

            if co.printL :
                print ("TablesNames", names)
                for fld in self.Flds :  fld.Mprint()

            for fld in self.Flds :
                if fld.src_name == '*' :
                    for na in names : self.Flds.append ( Field ( na, na ) )
                    self.Flds.append ( Field ( 'ROWNUM', 'ROWNUM' ) )
                elif fld.src_name == 'ROWNUM' : fld.src_num = -1
                else :    
                    fld.src_num = names.index(fld.src_name)
                    if fld.src_num == -1 :
                        print ("No Column for", fld.src_name, "*****************************")
                        exit (-1)    
            self.KillField ('*')
            for fld in self.Flds : fld.Mprint()

            self.NoC = len(self.Flds)
            maxNoR = 50000;
            for fld in self.Flds : fld.tb = zeros ( maxNoR, float64 ) 

            NoR = 0;
            if self.FileType != 'mtr2' and self.FileType != 'matr2' :                   #  tbl
              for line in fi :
                nums_row = line.split()
           #     print ( 'F', nums_row[0] )
                if len(nums_row) == 0: continue
                if nums_row[0] == '#END#':  break
                if nums_row[0][0] == '#'    :  continue
                wher = where[0:]
                OK = True
                for fld in self.Flds :
                    if fld.src_num == -1 :   fld.tb[NoR] = NoR
                    else:
                        fld.tb[NoR] = floatGradNaN( nums_row[fld.src_num] )
#                        try :
 #                           fld.tb[NoR] = float ( nums_row[fld.src_num] )
  #                      except :
   #                         gra = nums_row[fld.src_num].index('�')
    #                        if gra >=0 :
     #                           fld.tb[NoR] = float (nums_row[fld.src_num][:gra]) + float(nums_row[fld.src_num][gra+1:])/60.
      #     #                     print (fld.tb[NoR], NoR)
       #                     else :
        #                        fld.tb[NoR] = NaN
                        if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
                if not OK: continue
                if len (where) >= 3 :                       #check  where
                    wher = where.replace('ROWNUM',str(NoR))
                    for fld in self.Flds :
#                        print 'fld.name', fld.name
                        wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
 #                       print 'wher', wher
                    if eval (wher, {}) == False : continue    

                NoR += 1
                if NoR >= maxNoR :                          # resize
                    maxNoR *= 2
                    print ('resize to',  maxNoR)
                    for fld in self.Flds : fld.tb = resize (fld.tb, maxNoR ) 
            else :                                                          # == 'matr2'
              XX = fi.readline().split()                      # x1
#              print len(XX), XX
              for line in fi :
                nums_row = line.split()
                if len (nums_row) == 0: continue
                wher = where[0:]
                for col, num in enumerate(nums_row) :
                    OK = True
                    if col == 0 : Y = float (num);  continue

                    for fld in self.Flds :
                        if   fld.src_num == -1 :   fld.tb[NoR] = NoR
                        elif fld.src_num ==  0 :   fld.tb[NoR] = float( XX[col-1] )  
                        elif fld.src_num ==  1 :   fld.tb[NoR] = Y
                        else:  # fld.src_num ==  2
                            try :
                                fld.tb[NoR] = float ( num )
                            except :
                                fld.tb[NoR] = NaN
                            if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
                    if not OK: continue
                    if len (where) >= 3 :                       #check  where
                      wher = where.replace('ROWNUM',str(NoR))
                      for fld in self.Flds:
                        wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
 #                       for c in range(self.NoC):
    #                        wher = SubstitudeName(wher, self.cols[c], str(tbl[NoR, c]))

                      if eval (wher,{}) == False : continue
                    NoR += 1
                    if NoR >= maxNoR :                          # resize
                        maxNoR *= 2
                        print ('resize to',  maxNoR)
                        for fld in self.Flds : fld.tb = resize (fld.tb, maxNoR ) 
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR ) 
#            print self.Flds[2].tb
            self.NoR = NoR
                
    def Read27_ASC( self, where ) :            # == 'matr Grid'               #    tbl -> fld.tb
        with open( self.fromFile, "r") as fi:
            names = ['X','Y','Z']
            if co.printL :  print ("TablesNames", names)
            for c in self.Flds :
                try :
#                    col_num = names.index(c.src_name) #[0])
                    c.src_num = names.index(c.src_name) #[0])
                except :    
                        print ("No Column for", c.src_name, "*****************************")
                        return    
            for fld in self.Flds : fld.Mprint()


            grdX      = int(fi.readline().split()[1])
            grdY      = int(fi.readline().split()[1])
            XLLCORNER = float(fi.readline().split()[1])
            YLLCORNER = float(fi.readline().split()[1])
            CELLSIZE  = float(fi.readline().split()[1])
            NDT  = float(fi.readline().split()[1])
            if co.printL: print ("ReadGrig from", self.fromFile)  ###, self.cols  ###, self.nums
            if co.printL: print (grdX, grdY, XLLCORNER, YLLCORNER, CELLSIZE, NDT)

            x1 = zeros(grdX, float64)
            x2 = zeros(grdY, float64)
            for i in range(grdX): x1[i] = XLLCORNER + CELLSIZE * (i + 0.5)
            for j in range(grdY): x2[j] = YLLCORNER + CELLSIZE * (grdY-1 - j + 0.5)

###            self.NoC = len(self.cols)
            self.NoC = len(self.Flds)
            maxNoR = 50000;
            tbl = zeros ( (maxNoR, self.NoC), float64 )
            NoR = 0;

            r = where.split ('XYin')
            if len(r) == 2:  Rect = readListFloat19 (r[1])
            else          :  Rect = []
            if len (Rect) == 4 :
                  xmi = int ( ceil( (Rect[0] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10 ) )    # !!!!!!!!!!!
                  xma = int (floor( (Rect[2] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10 ) )    # !!!!!!!!!!!
                  yma = grdY - 1 - int ( ceil( (Rect[1] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10 ))
                  ymi = grdY - 1 - int (floor( (Rect[3] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10 ))
            else :
                  xmi = 0
                  xma = grdX - 1
                  yma = grdY - 1
                  ymi = 0
                  print ("|||  xmi, xma, ymi, yma", xmi, xma, ymi, yma)
            
            for s in range(ymi) : fi.readline()
            line_num = ymi

            for s in range(yma-ymi+1):
                in_num = fi.readline().split()
                for col in range (len(x1)) :
                  OK = True                
                  if col < xmi : continue
                  if col > xma : break
                  for c in range(self.NoC) :
###                    if   self.nums[c]== 0 : tbl[NoR,c] = float(x1[col])
   ###                 elif self.nums[c]== 1 : tbl[NoR,c] = float(x2[line_num])
                    if   self.Flds[c].src_num== 0 : tbl[NoR,c] = float(x1[col])
                    elif self.Flds[c].src_num== 1 : tbl[NoR,c] = float(x2[line_num])
                    else :
                                            tbl[NoR,c] = float(in_num[col])
                                            if tbl[NoR,c] == NDT: tbl[NoR,c] = NaN
 #                                               in_num[col]= NaN
 #                                           tbl[NoR,c] = float(in_num[col])
                                            if not self.useNaN and isnan(tbl[NoR,c]) : OK = False; break   
                  if not OK: continue
#                  if len (where) >= 3 :                       #check  where
 #                   wher = where.replace('ROWNUM',str(NoR))   #  �������� �� copy
  #                  for c in range(self.NoC) :
   #                     wher = SubstitudeName ( wher, self.cols[c], str(tbl[NoR,c]) )
    #                if eval (wher,{}) == False : continue    
                  NoR += 1
                  if NoR >= maxNoR :                          # resize
                    maxNoR *= 2
                    print ('resize to',  maxNoR)
                    tbl = resize (tbl, (maxNoR, self.NoC) )
                line_num += 1    
###            self.tbl = resize (tbl, (NoR, self.NoC) )
            tbl = resize (tbl, (NoR, self.NoC) )
            for ifld, fld in enumerate(self.Flds): fld.tb = tbl[:, ifld]  ############### Kill ######
#            for r in range (100) :
 #               print i, tbl[i,0], tbl[i,1], tbl[i,2]
            self.NoR = NoR
            return

    def WriteSvFtbl ( self, OutName, printL=0 ) :
        with open(OutName,'w') as f:  
            for fld in self.Flds :  f.write ( fld.name + '  ' )
            f.write ( '#SvFver_62_tbl' )
            for r in range ( self.NoR ) :
                f.write ('\n' )
                for fld in self.Flds :  f.write ('\t'+str(fld.tb[r]) )
        if printL : print ("Write to", OutName)

    def Old_Read19_ASC(self, _cols, where):  # == 'matr Grid'               #    tbl -> fld.tb
        with open(self.fromFile, "r") as fi:
            names = ['X', 'Y', 'Z']
            if co.printL:  print ("TablesNames", names)
            for c in _cols:
                try:
                    col_num = names.index(c[0])
                except:
                    print ("No Column for", c[0], "*****************************")
                    return
                self.nums.append(col_num)
                if len(c) == 3: c[0] = c[2]  ##  ????????????

            for i in range(len(_cols)):       self.cols.append(_cols[i][0])

            grdX = int(fi.readline().split()[1])
            grdY = int(fi.readline().split()[1])
            XLLCORNER = float(fi.readline().split()[1])
            YLLCORNER = float(fi.readline().split()[1])
            CELLSIZE = float(fi.readline().split()[1])
            NDT = float(fi.readline().split()[1])
            if co.printL: print ("ReadGrig from", self.fromFile, self.cols, self.nums)
            if co.printL: print (grdX, grdY, XLLCORNER, YLLCORNER, CELLSIZE, NDT)

            x1 = zeros(grdX, float64)
            x2 = zeros(grdY, float64)
            for i in range(grdX): x1[i] = XLLCORNER + CELLSIZE * (i + 0.5)
            for j in range(grdY): x2[j] = YLLCORNER + CELLSIZE * (grdY - 1 - j + 0.5)

            self.NoC = len(self.cols)
            maxNoR = 50000;
            tbl = zeros((maxNoR, self.NoC), float64)
            NoR = 0;

            r = where.split('XYin')
            if len(r) == 2:
                Rect = readListFloat19(r[1])
            else:
                Rect = []
            if len(Rect) == 4:
                xmi = int(ceil((Rect[0] - XLLCORNER - CELLSIZE / 2) / CELLSIZE - 1e-10))  # !!!!!!!!!!!
                xma = int(floor((Rect[2] - XLLCORNER - CELLSIZE / 2) / CELLSIZE + 1e-10))  # !!!!!!!!!!!
                yma = grdY - 1 - int(ceil((Rect[1] - YLLCORNER - CELLSIZE / 2) / CELLSIZE - 1e-10))
                ymi = grdY - 1 - int(floor((Rect[3] - YLLCORNER - CELLSIZE / 2) / CELLSIZE + 1e-10))
            else:
                xmi = 0
                xma = grdX - 1
                yma = grdY - 1
                ymi = 0
                print ("|||  xmi, xma, ymi, yma", xmi, xma, ymi, yma)

            for s in range(ymi): fi.readline()
            line_num = ymi

            for s in range(yma - ymi + 1):
                in_num = fi.readline().split()
                for col in range(len(x1)):
                    OK = True
                    if col < xmi: continue
                    if col > xma: break
                    for c in range(self.NoC):
                        if self.nums[c] == 0:
                            tbl[NoR, c] = float(x1[col])
                        elif self.nums[c] == 1:
                            tbl[NoR, c] = float(x2[line_num])
                        else:
                            tbl[NoR, c] = float(in_num[col])
                            if not self.useNaN and isnan(tbl[NoR, c]): OK = False; break
                    if not OK: continue
                    #                  if len (where) >= 3 :                       #check  where
                    #                   wher = where.replace('ROWNUM',str(NoR))   #  �������� �� copy
                    #                  for c in range(self.NoC) :
                    #                     wher = SubstitudeName ( wher, self.cols[c], str(tbl[NoR,c]) )
                    #                if eval (wher,{}) == False : continue
                    NoR += 1
                    if NoR >= maxNoR:  # resize
                        maxNoR *= 2
                        print ('resize to', maxNoR)
                        tbl = resize(tbl, (maxNoR, self.NoC))
                line_num += 1
            self.tbl = resize(tbl, (NoR, self.NoC))
            #            self.tb_cols = names
            self.NoR = NoR
            return


